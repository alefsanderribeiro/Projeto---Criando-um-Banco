from pathlib import Path
import sys
sys.path.append(str(Path().absolute()))
import openpyxl
from banco_de_dados.banco_dados import arquivo_contas
from cryptography.fernet import Fernet
import argon2

ARQUIVO_CHAVE_CRYPTO = Path() / "autenticacao" / "chave.txt"

argon2Hasher = argon2.PasswordHasher(
    time_cost=3, # número de interações 
    memory_cost=256 * 1024, # 256mb
    parallelism=1, # how many parallel threads to use
    hash_len=64, # the size of the derived key
    salt_len=32 # the size of the random generated salt in bytes
)

password = b"1612772310"

hash_encoded = argon2.PasswordHasher(hash_len=64, 
                                     salt_len=32, 
                                     time_cost=3, 
                                     memory_cost=256 * 1024, 
                                     parallelism=1).hash(password)

# chave = Fernet.generate_key()
# print("Chave AES gerado: ", chave)
# f = Fernet(chave)
# hash_criptografado = f.encrypt(hash_value.encode())
# print(hash_criptografado)   



# ph = argon2.PasswordHasher()
# # arg = argon2Hasher
# password = b"1612772310"
 
# hash = argon2Hasher.hash(password).encode('ascii')

# print(hash)
# password2 = "1612772310"
# try:
#     verifyValid = argon2Hasher.verify(hash, password2)
#     print("Senha correta")
# except argon2.exceptions.VerifyMismatchError:
#     print("Senha inválida")



# class Senhaaa: 
    
#     def __init__(self, numero_conta):
#         self.numero_conta = numero_conta
    

#     def _verificar_hash_senha(self):
#         reHash = argon2Hasher.check_needs_rehash(hash)
#         if reHash:
#             novo_para = argon2.Parameters(type, version, salt_len, hash_len, time_cost, memory_cost, parallelism)

    
    
#     def _verificar_senha(self):
        
#         senha_atual = str(input("Digita a sua senha..."))
#         self.numero_conta = int(self.numero_conta)
#         f = openpyxl.load_workbook(arquivo_contas)
#         aba = f.active
#         for i in range(1, len(aba['A']) + 1):
#             if self.numero_conta == aba[f'A{i}'].value:
#                 try:
#                     descriptografado = Crypt().descrypt(aba[f'H{i}'].value)
#                     print(descriptografado)
#                     resultado = argon2Hasher.verify(descriptografado, senha_atual)
#                     resultado = True
#                     print(resultado)
#                     print("Senha correta")
#                     break
#                 except:
#                     resultado = False

#         return resultado

#     def _inserir_senha_nova(self, nova_senha):
#         hash = argon2Hasher.hash(nova_senha)
#         hash = hash.split("$")
#         hash = hash[-1]
#         hash_argon2_cript = Crypt().crypt(hash)
        
#         f = openpyxl.load_workbook(arquivo_contas)
#         aba = f.active

#         for i in range(1, len(aba['A']) + 1):
#             if self.numero_conta == aba[f'A{i}'].value:
#                 senha_nova = aba.cell(row = i, column = 8).value = hash_argon2_cript
#                 print("Foi gravado a nova senha")
#                 f.save(arquivo_contas)
#                 break

#     def _nova_senha(self):
#         while True:
#             nova_senha1 = str(input("Digita a nova senha."))
#             nova_senha2 = str(input("Digita novamente a nova senha."))
            
#             if nova_senha1 == nova_senha2:
#                 resultado = self._inserir_senha_nova(nova_senha1)
#                 break
#             else:
#                 print("As senhas apresentada não são identicas, digite novamente!")
#         return nova_senha1

#     def alterar_senha(self):

#         while True:
            
#             senha = self._verificar_senha()
#             if senha == True:
#                 self._nova_senha()
#                 print("Senha alterada com sucesso!")
#                 break

#             else:
#                 print("Senha errada, digita novamente!")
    
        
# class Crypttt:
#     def __init__(self):
#         pass
    
#     def __gerar_chave(self, arquivo):
#         chave = Fernet.generate_key()
#         with open(arquivo,"wb") as arquivo_chave:
#             arquivo_chave.write(chave)
            
#     def __ler_chave(self, arquivo):
#         with open(arquivo,"rb") as arquivo_chave:
#             chave = arquivo_chave.read()
#         return chave

#     def chave_crypto(self):
#         try:
#             chave = self.__ler_chave(ARQUIVO_CHAVE_CRYPTO)
#         except FileNotFoundError:
#             self.__gerar_chave(ARQUIVO_CHAVE_CRYPTO)
#             chave = self.__ler_chave(ARQUIVO_CHAVE_CRYPTO)
#         return chave

#     def crypt(self, crypt):
#         chave = self.chave_crypto()
#         f = Fernet(chave)
#         crypt = f.encrypt(crypt.encode())
#         return crypt
    
#     def descrypt(self, descrypt):
#         chave = self.chave_crypto()
#         f = Fernet(chave)
#         descrypt = f.decrypt(descrypt.encode())
#         return descrypt
        


# chave = Fernet.generate_key()
# print("Chave AES gerado: ", chave)
# f = Fernet(chave)
# hash_criptografado = f.encrypt(hash_value.encode())
# print(hash_criptografado)   