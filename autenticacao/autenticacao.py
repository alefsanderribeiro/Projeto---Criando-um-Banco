from pathlib import Path
import sys
sys.path.append(str(Path().absolute()))
import openpyxl
from banco_de_dados.banco_dados import arquivo_clientes, arquivo_contas
from contas import cliente
   
def cadastro(CPF):

    f = openpyxl.load_workbook(arquivo_clientes)
    aba = f.active
    
    for i in range(1, len(aba['C']) + 1):
        
        if CPF == aba[f'C{i}'].value:
            resultado = True
            break
        else:
            resultado = False
    return resultado



def _CONTA(CONTA):
    
    CONTA = int(CONTA)
    f = openpyxl.load_workbook(arquivo_contas)
    aba = f.active
    contas = ""

    for i in range(1, len(aba['A']) + 1):
        
        if CONTA == aba[f'A{i}'].value:
            agencia = aba[f'B{i}'].value
            conta = aba[f'A{i}'].value
            tipo_conta = aba[f'D{i}'].value
            status_conta = aba[f'E{i}'].value
            dados = f"Agencia: {agencia} - Conta: {conta} - Tipo de Conta: {tipo_conta} - Status da Conta: {status_conta}\n"
            contas += dados
            
        else:
            pass
    return contas
    

def _contas_no_CPF(CPF):
    
    f = openpyxl.load_workbook(arquivo_contas)
    aba = f.active
    ID = cliente.Cliente(CPF)._informa_ID()
    contas = ""

    for i in range(1, len(aba['A']) + 1):
        
        if ID == aba[f'C{i}'].value:
            agencia = aba[f'B{i}'].value
            conta = aba[f'A{i}'].value
            tipo_conta = aba[f'D{i}'].value
            status_conta = aba[f'E{i}'].value
            dados = f"Conta: {conta} - Agencia: {agencia} - Tipo de Conta: {tipo_conta} - Status da Conta: {status_conta}\n"
            contas += dados
        else:
            pass
    return contas

def contas(cpf_or_conta):
    
    if 1 <= len(cpf_or_conta) <= 10:
        print(_CONTA(cpf_or_conta))
        
    elif len(cpf_or_conta) == 11:
        print(_contas_no_CPF(cpf_or_conta) )
            
    #return resultado