
import openpyxl
import pandas as pd


def verficar_cadastro(cpf_usuario):

    f = openpyxl.load_workbook("dados\clientes.xlsx")
    aba = f.active
        
    for i in range(1, len(aba['C']) + 1):
        
        if cpf_usuario == aba[f'C{i}'].value:
            resultado = True
            break
        else:
            resultado = False
    return resultado
            
            
def adicionar_usuario(dados):
    verif = verficar_cadastro(dados[0])
 
    if verif == False:
        f = openpyxl.load_workbook("dados\clientes.xlsx")
        aba = f.active
        line = len(aba['A']) + 1
        id_cadastro = aba.cell(row = line, column = 1).value = int(line - 1) #criado um ID para o cliente
        status_cadastro = aba.cell(row = line, column = 2).value = "Ativado" # colocando o Status do uusário de "Ativado"
        
        for i in range(len(dados)): #inserir os dados na mesma linha
            aba.cell(row = line, column = 3+i).value = dados[i]
        f.save("dados\clientes.xlsx")
        print(f"Cadastro realizado com Sucesso!\nID cliente n° {id_cadastro}")

    else:
        print(f"O CPF {dados[0]} já está cadastrado. Realize o cadastro de um novo CPF")


def desativar_usuario(cpf_do_cliente):
    
    f = openpyxl.load_workbook("dados\clientes.xlsx")
    aba = f.active
        
    for i in range(1, len(aba['C']) + 1):
        if cpf_do_cliente == aba[f'C{i}'].value:
            id_cadastro = aba[f'A{i}'].value
            status_cadastro = aba[f'B{i}'].value
            if status_cadastro != "Desativado":
                aba.cell(row = i, column = 2).value = "Desativado"
                print(f"O cadastro com o ID cliente n° {id_cadastro}, foi desativado com sucesso!")
                f.save("dados\clientes.xlsx")
                break
            else:
                print("O cadastro já está Desativado")
        else:
            print("CPF não encontrado.")

def ativar_usuario(cpf_do_cliente):
    
    f = openpyxl.load_workbook("dados\clientes.xlsx")
    aba = f.active
    result = "Não encontrado"
 
    for i in range(1, len(aba['C']) + 1):
        id_cadastro = aba[f'A{i}'].value
        status_cadastro = aba[f'B{i}'].value
        
        if cpf_do_cliente != aba[f'C{i}'].value:
            pass
        #continuar arrumando aqui - Continuar arrumando a parte de aparecer que não foi possível encontrar o cadastro.
        elif cpf_do_cliente == aba[f'C{i}'].value:            
            if status_cadastro != "Ativado":
                aba.cell(row = i, column = 2).value = "Ativado"
                f.save("dados\clientes.xlsx")
                result = f"O cadastro com o ID cliente n° {id_cadastro}, foi Ativado com sucesso!"
                break
            else:
                result = "O cadastro já está Ativado"
    return result


dados = ["01912841223", "Alefsander Ribeiro", "23/10/1994", "R. Escorpião, n° 11700 - Bairro: Ulisses Guimarães - Porto Velho/RO."]
dados1 = ["0191284122f1233", "Alefsander Ribeiro", "23/10/1994", "R. Escorpião, n° 11700 - Bairro: Ulisses Guimarães - Porto Velho/RO."]
dados2 = ["04240084245", "Geissilaine Verônica", "19/107/1998", "R. Maué, n° 4445 - Bairro: Socialista - Porto Velho/RO."]
dados3 = ["123461365123", "Alefsander Ribeiro", "23/10/1994", "R. Escorpião, n° 11700 - Bairro: Ulisses Guimarães - Porto Velho/RO."]
dados4 = ["123461365", "Alefsander Ribeiro", "23/10/1994", "R. Escorpião, n° 11700 - Bairro: Ulisses Guimarães - Porto Velho/RO."]

#adicionar_usuario(dados4)

ativ = ativar_usuario("019128423")

print(ativ)
#desativar_usuario("019128233")
#desativar_usuario("123461365123")


f = pd.read_excel("dados\clientes.xlsx")
print(f)


""" f = openpyxl.load_workbook("dados\clientes.xlsx")
aba_ativa = f.active
print(len(aba_ativa['A']))
print(aba_ativa.iter_rows())
aba_ativa.delete_rows(3)
f.save("dados\clientes.xlsx")




def ajuste_dados():
  
    f = openpyxl.load_workbook("dados\clientes.xlsx")
    aba = f['clientes']
    
    
    for row in aba.iter_rows():

        if not all(cell.value for cell in row): 
        
            aba.delete_rows(row[0].row, 1) 
            ajuste_dados() 
            
    aba.save("dados\clientes.xlsx")
    
    
#ajuste_dados()
    if __name__ == '__main__': 

        for row in sheet: 
            remove(sheet) 
            
        print("Maximum rows after removing:",sheet.max_row) 
    
        path = './openpy.xlsx'
        book.save(path) 
"""