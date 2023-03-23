from pathlib import Path
import sys
sys.path.append(str(Path().absolute()))
import openpyxl
import pandas as pd
import autenticacao.autenticacao_conta  as aut_conta
import autenticacao.autenticacao_senha  as aut_senha
from autenticacao.criptografia  import Crypt
import banco_de_dados.banco_dados as bd


class Cliente:
    def __init__(self, CPF: str):
        self.CPF = Crypt().crypt(CPF)

    
    def __str__(self):
        return print(f"Dados do usuário:\
            CPF: {self.CPF}")
    
    def __doc__(self):
        pass
         
    def adicionar_usuario(self):
        
        verif = bd.Procurar.cpf(self.CPF)
        if verif:
            print(f"O CPF {self.CPF} já está cadastrado. Realize o cadastro de um novo CPF")
        
        else:
        
            nome = Crypt().crypt(str(input("Digita o seu nome completo:\n ")))
            data_nascimento = Crypt().crypt(str(input("Digita a sua data de Nascimento:\n ")))
            endereçoRua = Crypt().crypt(str(input("Vamos cadastrar o seu endereço:\nDigita o nome da sua rua/av:\n")))
            endereçoNumero = Crypt().crypt(str(input("Digita o nome número do seu endereço:\n")))
            endereçoBairro = Crypt().crypt(str(input("Digita o bairro:\n")))
            endereçoCidade = Crypt().crypt(str(input("Digita o nome da cidade:\n")))
            endereçoEstado = Crypt().crypt(str(input("Digita o nome do Estado:\n")))

            telefone = Crypt().crypt(str(input("Digita o seu telefone pessoal:\n")))
            email = Crypt().crypt(str(input("Digita o seu email pessoal:\n")))
            # dados_completo = [nome, data_nascimento, 
            #         f"{endereçoRua}, {endereçoNumero} - bairro: {endereçoBairro} - {endereçoCidade}", 
            #         telefone, email]
            
            senha = aut_senha.Senha(self.CPF)._nova_senha(conta_ou_usuario="usuario")

            resultado = bd.Inserir.dados_cliente(self.CPF, 
                                                nome=nome, 
                                                data_nascimento=data_nascimento, 
                                                enderecoRua=endereçoRua,
                                                enderecoNumero=endereçoNumero,
                                                enderecoBairro=endereçoBairro,
                                                enderecoCidade=endereçoCidade,
                                                enderecoEstado=endereçoEstado,
                                                telefone=telefone, 
                                                email=email, 
                                                senha=senha)
            
            if resultado:

                id_cadastro = bd.Procurar(self.CPF).procurar_id()
                status_cadastro = bd.Procurar(self.CPF).procurar_status()
                print(f"Cadastro realizado com Sucesso!\nID cliente n° {id_cadastro}, \
                    com o status de {status_cadastro}.")
            else:
                print("Não foi possível cadastrar o usuário!")


    def desativar_usuario(self):
        
        f = openpyxl.load_workbook(bd.arquivo_clientes)
        aba = f.active
        
        for i in range(1, len(aba['C']) + 1):
            if self.CPF == aba[f'C{i}'].value:
                id_cadastro = aba[f'A{i}'].value
                status_cadastro = aba[f'B{i}'].value
                if status_cadastro != "Desativado":
                    aba.cell(row = i, column = 2).value = "Desativado"
                    print(f"O cadastro com o ID cliente n° {id_cadastro}, foi desativado com sucesso!")
                    f.save(bd.arquivo_clientes)
                    break
                elif status_cadastro == "Desativado":
                    print("Cadastro já encontra-se desativado!")
                    
    def _informa_ID(self):
        f = openpyxl.load_workbook(bd.arquivo_clientes)
        aba = f.active
        
        for i in range(1, len(aba['C']) + 1):
            
            if self.CPF == aba[f'C{i}'].value:
                id_cadastro = aba[f'A{i}'].value

                break
        return id_cadastro
    
    def _informa_Status(self):
        f = openpyxl.load_workbook(bd.arquivo_clientes)
        aba = f.active
        
        for i in range(1, len(aba['C']) + 1):
            
            if self.CPF == aba[f'C{i}'].value:
                status_cadastro = aba[f'B{i}'].value
                resultado = status_cadastro
                break
        return resultado

    
    def ativar_usuario(self):
        
        f = openpyxl.load_workbook(bd.arquivo_clientes)
        aba = f.active
    
        for i in range(1, len(aba['C']) + 1):

            if self.CPF == aba[f'C{i}'].value:  
                id_cadastro = aba[f'A{i}'].value
                status_cadastro = aba[f'B{i}'].value
       
                if status_cadastro == "Desativado":
                    aba.cell(row = i, column = 2).value = "Ativado"
                    f.save(bd.arquivo_clientes)
                    print(f"O cadastro com o ID cliente n° {id_cadastro}, foi Ativado com sucesso!")
                    break
                elif status_cadastro == "Ativado":
                    print("Cadastro já encontra-se desativado!")
                    
def listar_usuarios():
    f = pd.read_excel(bd.arquivo_clientes)
    print(f)
      