# Nesse arquivo ficará todas as funções bancárias e configurações do banco
from datetime import datetime
import openpyxl
import pandas as pd
from pathlib import Path
from contas import cliente as fu
ARQUIVO_CONTAS = Path() / "contas" / "dados" / "contas.xlsx"

class banco:
    
    def __init__(self):
        self.agencia = "0001"
        self.conta = "conta"
        self.saldo = "saldo"
        self.qtd_saque = "quantidade_saque_efetuado"
        self.historico_deposito = "historico_deposito"
        self.historico_saque = "historico_saque"
        self.LIMITE = {"Valor_Saque": 500, "Qtd_Saque": 3}
        self.LIMITE_VALOR_SAQUE = 500
        self.LIMITE_QUANT_SAQUE = 3
        
    def __str__(self):
        
        return "Este é o meu banco."
    
    
    def _LIMITES_CONTA(self):
        pass
        
    def _aumentar_limite(self):
        pass
    
    
    def _diminuir_limite(self):
        pass        


    def _verif_conta_CONTA(self, CONTA):
        
        CONTA = int(CONTA)
        f = openpyxl.load_workbook(ARQUIVO_CONTAS)
        aba = f.active
        contas = ""

        for i in range(1, len(aba['A']) + 1):
            
            if CONTA == aba[f'A{i}'].value:
                agencia = aba[f'B{i}'].value
                conta = aba[f'A{i}'].value
                tipo_conta = aba[f'D{i}'].value
                status_conta = aba[f'E{i}'].value
                dados = f"Agencia: {agencia} - Conta: {conta} - Tipo de Conta: {tipo_conta} - Status da Conta: {status_conta}\n"
                contas += dados
                
            else:
                pass
        return contas
    

    def _verif_contas_CPF(self, CPF):
        
        f = openpyxl.load_workbook(ARQUIVO_CONTAS)
        aba = f.active
        ID = fu.usuario(CPF)._informa_ID()
        contas = ""

        for i in range(1, len(aba['A']) + 1):
            
            if ID == aba[f'C{i}'].value:
                agencia = aba[f'B{i}'].value
                conta = aba[f'A{i}'].value
                tipo_conta = aba[f'D{i}'].value
                status_conta = aba[f'E{i}'].value
                dados = f"Conta: {conta} - Agencia: {agencia} - Tipo de Conta: {tipo_conta} - Status da Conta: {status_conta}\n"
                contas += dados
            else:
                pass
        return contas
    
    def verificar_contas(self, cpf_or_conta):
        
        if 1 <= len(cpf_or_conta) <= 10:
            print(self._verif_conta_CONTA(cpf_or_conta))
            
        elif len(cpf_or_conta) == 11:
            print(self._verif_contas_CPF(cpf_or_conta) )
               
        #return resultado
        

    def _verificar_senha(self, numero_conta):
        
        senha_atual = str(input("Digita a sua senha..."))
        numero_conta = int(numero_conta)
        f = openpyxl.load_workbook(ARQUIVO_CONTAS)
        aba = f.active
            
        for i in range(1, len(aba['A']) + 1):

            if numero_conta == aba[f'A{i}'].value and senha_atual == aba[f'H{i}'].value:
                resultado = True
                break
            elif numero_conta != aba[f'A{i+1}'].value:

                resultado = False
                
        return resultado
    
    def _inserir_senha_nova(self, numero_conta, nova_senha):
        
        f = openpyxl.load_workbook(ARQUIVO_CONTAS)
        aba = f.active
        resultado = ""
        for i in range(1, len(aba['A']) + 1):
            if numero_conta == aba[f'A{i}'].value:
                senha_nova = aba.cell(row = i, column = 8).value = nova_senha
                resultado = "Foi gravado a nova senha"
                f.save(ARQUIVO_CONTAS)
                break
            else:
                resultado = "Não foi gravado a nova senha"
        return resultado
        
    
    
    def _nova_senha(self, numero_conta):
        while True:
            nova_senha1 = str(input("Digita a nova senha."))
            nova_senha2 = str(input("Digita novamente a nova senha."))
            
            if nova_senha1 == nova_senha2:
                self._inserir_senha_nova(numero_conta, nova_senha1)
                break
            else:
                print("As senhas apresentada não são identicas, digite novamente!")
        return nova_senha1
    
    def alterar_senha(self, numero_conta):
    
        while True:
            
            senha = self._verificar_senha(numero_conta)
            if senha == True:
                self._nova_senha(numero_conta)
                print("Senha alterada com sucesso!")
                break

            else:
                print("Senha errada, digita novamente!")
                
    
    def criar_conta(self, CPF):
        ID = fu.usuario(CPF)._informa_ID()
        status = fu.usuario(CPF)._informa_Status()
        
        if status == "Desativado":
            resultado = "O usuário está desativado!"
        elif status == "Ativado":
            
            f = openpyxl.load_workbook(ARQUIVO_CONTAS)
            aba = f.active
    
            tipo_conta = self._tipo_conta()

            line = len(aba['A']) + 1
            numero_conta = aba.cell(row = line, column = 1).value = int(line - 1) # criado o número da conta
            agencia = aba.cell(row = line, column = 2).value = self.agencia # inserindo o número da agencia
            id_usuario = aba.cell(row = line, column = 3).value = ID
            t_conta = aba.cell(row = line, column = 4).value = tipo_conta # inserindo o Tipo de Conta no cadastro.
            status_conta = aba.cell(row = line, column = 5).value = "Ativado" # inserindo o Status de Conta no cadastro.
            limite_valor_saque = aba.cell(row = line, column = 6).value = self.LIMITE_VALOR_SAQUE # inserindo o valor limite do saque.
            limite_quant_saque = aba.cell(row = line, column = 7).value = self.LIMITE_QUANT_SAQUE # inserindo a quantidade do saque diário.
            f.save(ARQUIVO_CONTAS)
            senha = self._nova_senha(numero_conta)
            
            resultado = f"""Cadastro realizado com Sucesso!
                            Conta n°: {numero_conta}
                            Agência n°: {agencia}
                            CPF usuário n°: {CPF} (ID: {id_usuario})
                            Tipo de Conta: {t_conta}
                            Status da Conta: {status_conta}
                            Valor Limite por Saque: {limite_valor_saque}
                            Limite da Quantidade deSaque: {limite_quant_saque}
                            senha: {senha}
                            """
        return resultado




    def excluir_conta(self):
        pass
    
    
    
    def ativar_conta(self):
        pass
    
    
    
    def desativar_conta(self):
        pass
    
    
    
    def _tipo_conta(self):
        
        tipo_conta = ""
        
        while True:
            tipo_conta = str(input("""Informa o tipo de conta:
                                    'CC' = Conta Corrente
                                    'CP' = Conta Poupança
                                    'CS' = Conta Salário\n
                                    """)).upper()
        
            if tipo_conta == "CC":
                resultado = "CC"
                break

            elif tipo_conta == "CP":
                resultado = "CP"
                break

            elif tipo_conta == "CS":
                resultado = "CS"
                break
            else:
                print("Opção incorreta, insere novamente o Tipo de Conta!")
        return resultado
    
    def transferir_mesmo_banco(self):
        pass
    
    
        
    def deposito(self):

        while True:
            valor_deposito = float(input("Insere o valor que deseja depositar:\n"))
            if valor_deposito > 0:
                self.saldo += valor_deposito
                agora = datetime.now().strftime('%d/%m/%Y %H:%M')
                self.historico_deposito += f"Depósito em {agora} - Valor R$ {valor_deposito:.2f}\n"
                print(F"Deposito no valor de {valor_deposito} feito com sucesso.\n")
                continuar = str(input("Deseja realizar mais um deposito? \n S ou N?")).upper()
                
                if continuar == "S":
                    continue
                elif continuar == "N":
                    break   
            else:
                print("Valor não permitido.")
        return self.saldo
        
    def saque(self):
        while True:
            valor_saque = float(input("Insere o valor que deseja sacar:\n"))
            if 0 < valor_saque <= self.LIMITE["Valor_Saque"] and self.qtd_saque < self.LIMITE["Qtd_Saque"]\
                and valor_saque <= self.saldo:
                self.saldo -= valor_saque
                agora = datetime.now().strftime('%d/%m/%Y %H:%M')
                self.historico_saque += f"Saque em {agora} - Valor R$ {valor_saque:.2f}\n"
                self.qtd_saque += 1
                print(f"Saque no valor de {valor_saque} feito com sucesso.\n")
                continuar = str(input("Deseja realizar mais um saque? \n S ou N?")).upper()
                if continuar == "S":
                    continue
                elif continuar == "N":
                    break
            elif valor_saque > self.LIMITE["Valor_Saque"]:
                print(f'Valor maior do que o limite de saque: {self.LIMITE["Valor_Saque"]}')
            elif self.qtd_saque == self.LIMITE["Qtd_Saque"]:
                print("Não é possível efetuar o saque, pois atingiu o limite do saque")
                break
            elif self.saldo < valor_saque:
                print("Você não tem saldo para realizar o saque.\nPor favor, efetue mais depósitos")
            elif valor_saque <= 0:
                print("Valor não permitido.")
                
    def extrato(self):
        print ("**************** EXTRATO ****************")
        print("Não há lançamentos de Depósito." if self.historico_deposito == "" else self.historico_deposito)
        print("Não há lançamentos de Saque." if self.historico_saque =="" else self.historico_saque)
        print(f"Seu saldo atual é de R$ {self.saldo}\n")
        agora = datetime.now().strftime('%d/%m/%Y %H:%M')
        print(f"Data: {agora}")
        print("*****************************************")
        




        